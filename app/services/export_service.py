import uuid
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.game import Game
from app.models.reservation import Reservation


def build_whatsapp_message(
    game: Game,
    black_team: list[Reservation],
    white_team: list[Reservation],
    backups: list[Reservation],
) -> str:
    lines = [
        f"*{game.title}*",
        f"{game.game_date.strftime('%A, %d %B %Y')} "
        f"{game.start_time.strftime('%H:%M')}–{game.end_time.strftime('%H:%M')}",
        "",
        "\u2b1b *Team Black*",
    ]
    if black_team:
        for i, r in enumerate(black_team, 1):
            lines.append(f"{i}. {r.user.first_name} {r.user.last_name}")
    else:
        lines.append("(none)")

    lines += ["", "\u2b1c *Team White*"]
    if white_team:
        for i, r in enumerate(white_team, 1):
            lines.append(f"{i}. {r.user.first_name} {r.user.last_name}")
    else:
        lines.append("(none)")

    if backups:
        lines += ["", "\U0001f4cb *Backup list*"]
        for i, r in enumerate(backups, 1):
            lines.append(f"{i}. {r.user.first_name} {r.user.last_name}")

    return "\n".join(lines)


def build_excel_workbook(
    game: Game,
    black_team: list[Reservation],
    white_team: list[Reservation],
    backups: list[Reservation],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"

    headers = ["Team Black", "Team White", "Backups"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    max_rows = max(len(black_team), len(white_team), len(backups), 0)
    for i in range(max_rows):
        row = i + 2
        if i < len(black_team):
            r = black_team[i]
            ws.cell(row=row, column=1, value=f"{r.user.first_name} {r.user.last_name}")
        if i < len(white_team):
            r = white_team[i]
            ws.cell(row=row, column=2, value=f"{r.user.first_name} {r.user.last_name}")
        if i < len(backups):
            r = backups[i]
            ws.cell(row=row, column=3, value=f"{r.user.first_name} {r.user.last_name}")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
